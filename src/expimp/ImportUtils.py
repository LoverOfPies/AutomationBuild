from openpyxl import load_workbook

from src.db.DbUtils import add_multirow, add_row
from src.db.models.base.Prop import Prop
from src.db.models.base.Unit import Unit
from src.db.models.material.Material import Material
from src.db.models.material.MaterialCategory import MaterialCategory
from src.db.models.material.MaterialGroup import MaterialGroup
from src.db.models.material.MaterialProperty import MaterialProperty
from src.db.models.material.MaterialSubgroup import MaterialSubgroup
from src.gui.custom_uix.ErrorPopup import ErrorPopup


def import_single_row(filename, ui, list_name):
    wb = load_workbook(filename)
    try:
        sheet = wb.get_sheet_by_name(list_name)
    except KeyError:
        ErrorPopup(message='Неправильный файл').open()
        return
    values = []
    for cell in sheet['A']:
        if cell.row == 1:
            continue
        values.append([{'name': str(cell.value)}])
    data = dict([
        ('model_class', ui.model_class),
        ('value', values),
    ])
    add_multirow(data)
    ui.update_screen()


def import_material(filename, list_name):
    wb = load_workbook(filename)
    try:
        sheet = wb.get_sheet_by_name(list_name)
    except KeyError:
        ErrorPopup(message='Неправильный файл').open()
        return

    len_columns = len([row for row in sheet.columns])
    is_first = True

    material_category = None
    material_group = None
    material_subgroup = None
    material = None

    for row in sheet.rows:
        if is_first:
            is_first = False
            continue
        for col in range(len_columns):
            if col == 0 and row[col].value != None:
                data_material_category = dict([
                    ('model_class', MaterialCategory),
                    ('value', [{'name': str(row[col].value)}]),
                ])
                add_row(data_material_category)
                material_category = MaterialCategory.select().where(MaterialCategory.name == row[col].value)
            if col == 1 and row[col].value != None:
                value_material_group = [
                    {'name': str(row[col].value),
                     'material_category': material_category}
                ]
                data_material_group = dict([
                    ('model_class', MaterialGroup),
                    ('value', value_material_group),
                ])
                add_row(data_material_group)
                material_group = MaterialGroup.select().where(MaterialGroup.name == row[col].value)
            if col == 2 and row[col].value != None:
                value_material_subgroup = [
                    {'name': str(row[col].value),
                     'material_group': material_group}
                ]
                data_material_subgroup = dict([
                    ('model_class', MaterialSubgroup),
                    ('value', value_material_subgroup),
                ])
                add_row(data_material_subgroup)
                material_subgroup = MaterialSubgroup.select().where(MaterialSubgroup.name == row[col].value)
            if col == 3 and row[col].value != None:
                unit = Unit.select().where(Unit.name == row[4].value)
                value_material = [
                    {'name': str(row[col].value),
                     'articul': '0',
                     'unit': unit,
                     'subgroup': material_subgroup}
                ]
                data_material = dict([
                    ('model_class', Material),
                    ('value', value_material),
                ])
                add_row(data_material)
                material = Material.select().where(Material.name == row[col].value)
            if col == 5 and row[col].value != None:
                unit = Unit.select().where(Unit.name == row[7].value)
                prop = Prop.select().where(Prop.name == row[5].value)
                value_material = [
                    {'amount': str(row[6].value),
                     'material': material,
                     'prop': prop,
                     'unit': unit}
                ]
                data_material_property = dict([
                    ('model_class', MaterialProperty),
                    ('value', value_material),
                ])
                add_row(data_material_property)
